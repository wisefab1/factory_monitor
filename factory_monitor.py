"""
Factory Machine Monitor
=======================
1. Clear old CSV files
2. Download fresh CSV from web interface
3. Analyze last hour:
   - Program cycle times + averages
   - Downtime detection with reasons
   - Timeline chart per machine
   - 7-day history (SQLite)
   - Telegram alert if idle > 60 min
4. Open HTML report in browser
"""

import os
import time
import sqlite3
import csv
import sys
import json
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from collections import defaultdict

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, WebDriverException

# ── Configuration ─────────────────────────────────────────────────────────────
DOWNLOAD_DIR        = r"C:\Connectplan_raports"
CSV_FILE            = os.path.join(DOWNLOAD_DIR, "operation_history.csv")
MACHINING_RESULT    = os.path.join(DOWNLOAD_DIR, "machining_results.csv")
OUTPUT_HTML         = os.path.join(DOWNLOAD_DIR, "report.html")
DB_FILE             = os.path.join(DOWNLOAD_DIR, "history.db")
LOG_FILE            = os.path.join(DOWNLOAD_DIR, "factory_monitor.log")
URL                 = "http://192.168.1.210/csv/OutputCSVWeb.aspx?FactoryID=1&AreaID=1"
WAIT_TIME           = 300
HOURS_BACK          = 24
ALERT_THRESHOLD_MIN = 45

TELEGRAM_TOKEN   = "8474596481:AAEGyP1nB0vuRo4DkCLwzDbBXDV7Lab7lvU"
TELEGRAM_CHAT_ID = "656625394"

GITHUB_TOKEN = "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"
GITHUB_USER  = "wisefab1"
GITHUB_REPO  = "factory_monitor"
GITHUB_URL   = "https://wisefab1.github.io/factory_monitor/"

TARGET_TIME_FILE = r"\\wisefile\Wisefile\WF_Tootmine\Planeerimine\CNC toodete ajad pildid\CNC tehno.xlsm"

# ── Logging ───────────────────────────────────────────────────────────────────
def log(message: str):
    """Виводить повідомлення в консоль та зберігає у файл логів."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_line + "\n")
    except Exception as e:
        print(f"Failed to write log: {e}")

# ── Excel Target Time ─────────────────────────────────────────────────────────
# ── Excel Target Time ─────────────────────────────────────────────────────────
def normalize_program_name(name: str) -> str:
    """Нормалізує назву програми для порівняння
    
    Порядок обробки:
    1. Видалення розширення
    2. Визначення операції та видалення (op1-5, p1-5, -1 до -5)
    3. Видалення останньої L або R
    4. Переведення у верхній регістр
    5. Видалення пробілів, дефісів, підкреслень
    6. Заміна 101 на 100
    
    Приклад: "WF861-100L-P2.MIN" → "WF861100"
    Приклад: "WF080-920-2.MIN" → "WF080920"
    """
    if not name:
        return ""
    
    # 1. Видаляємо розширення
    if '.' in name:
        name = name.rsplit('.', 1)[0]
    
    # 2. Визначаємо та видаляємо номер операції з кінця
    # Спочатку шукаємо op1-5 або p1-5
    name_lower = name.lower()
    found_op = False
    for op in ['op5', 'p5', 'op4', 'p4', 'op3', 'p3', 'op2', 'p2', 'op1', 'p1']:
        for variant in [f'-{op}', f'_{op}', op]:
            if name_lower.endswith(variant):
                name = name[:len(name) - len(variant)]
                found_op = True
                break
        if found_op:
            break
    
    # Якщо не знайшли op/p, шукаємо просто -1, -2, -3, -4, -5 в кінці
    if not found_op:
        for digit in ['5', '4', '3', '2', '1']:
            if name.endswith(f'-{digit}') or name.endswith(f'_{digit}'):
                name = name[:-2]
                break
    
    # 3. Видаляємо останню L або R якщо вона є
    if name and (name[-1].upper() == 'L' or name[-1].upper() == 'R'):
        name = name[:-1]
    
    # 4. Переводимо в верхній регістр
    name = name.upper()
    
    # 5. Видаляємо пробіли, дефіси, підкреслення
    name = name.replace(' ', '').replace('-', '').replace('_', '')
    
    # 6. Замінюємо 101 на 100 (WF861-100L та WF861_101L - одна деталь)
    name = name.replace('101', '100')
    
    return name

def parse_program_name(program_name: str) -> tuple:
    """Розбирає назву програми на базову назву та номер операції
    
    Args:
        program_name: повна назва програми (напр. "WF861-100L-P3.MIN")
    
    Returns:
        tuple: (normalized_base, operation_number)
        Приклад: "WF861-100L-P3.MIN" → ("WF861100", 3)
                 "WF080-920-2.MIN" → ("WF080920", 2)
    """
    normalized = normalize_program_name(program_name)  # "WF861100"
    operation = get_operation_number(program_name)      # 3
    return (normalized, operation)

def get_operation_number(program_name: str) -> int:
    """Визначає номер операції з назви програми
    
    Номер операції завжди перед крапкою що йде перед розширенням файлу.
    Наприклад: WF861-100L-P2.MIN → OP2
               WF861-100L.MIN → OP1 (немає номера)
               PART-OP3.PRG → OP3
    """
    if not program_name or program_name == "—":
        return 1
    
    # Видаляємо розширення (все після останньої крапки)
    if '.' in program_name:
        base_name = program_name.rsplit('.', 1)[0]
    else:
        base_name = program_name
    
    # Переводимо в нижній регістр для перевірки
    base_lower = base_name.lower()
    
    # Шукаємо op5/p5/-5 в кінці назви
    if base_lower.endswith('op5') or base_lower.endswith('p5') or base_lower.endswith('-p5') or base_lower.endswith('-5') or base_lower.endswith('_5'):
        return 5
    # Шукаємо op4/p4/-4
    elif base_lower.endswith('op4') or base_lower.endswith('p4') or base_lower.endswith('-p4') or base_lower.endswith('-4') or base_lower.endswith('_4'):
        return 4
    # Шукаємо op3/p3/-3
    elif base_lower.endswith('op3') or base_lower.endswith('p3') or base_lower.endswith('-p3') or base_lower.endswith('-3') or base_lower.endswith('_3'):
        return 3
    # Шукаємо op2/p2/-2
    elif base_lower.endswith('op2') or base_lower.endswith('p2') or base_lower.endswith('-p2') or base_lower.endswith('-2') or base_lower.endswith('_2'):
        return 2
    # Шукаємо op1/p1/-1
    elif base_lower.endswith('op1') or base_lower.endswith('p1') or base_lower.endswith('-p1') or base_lower.endswith('-1') or base_lower.endswith('_1'):
        return 1
    # Якщо не знайдено жодного номера - це OP1
    else:
        return 1

def load_target_times():
    """Завантажує Target Time з Excel файлу на сервері
    
    Структура Excel (вкладка "Tehnoloogiad"):
    A - Назва програми
    K - Станок OP1, N - Час OP1
    P - Станок OP2, S - Час OP2
    U - Станок OP3, X - Час OP3
    Z - Станок OP4, AC - Час OP4
    AE - Станок OP5, AH - Час OP5
    """
    try:
        import openpyxl
        log(f"Loading target times from: {TARGET_TIME_FILE}")
        
        # Спроба 1: Прямий доступ до UNC шляху
        try:
            log("Attempting direct UNC path access...")
            wb = openpyxl.load_workbook(TARGET_TIME_FILE, read_only=True, data_only=True)
            log("✓ Direct UNC access successful")
        except Exception as e1:
            log(f"Direct UNC access failed: {e1}")
            
            # Спроба 2: Через pathlib
            try:
                from pathlib import Path
                log("Attempting pathlib access...")
                path = Path(TARGET_TIME_FILE)
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                log("✓ Pathlib access successful")
            except Exception as e2:
                log(f"Pathlib access failed: {e2}")
                
                # Спроба 3: Через os.path.normpath
                try:
                    log("Attempting normalized path access...")
                    normalized = os.path.normpath(TARGET_TIME_FILE)
                    log(f"Normalized path: {normalized}")
                    wb = openpyxl.load_workbook(normalized, read_only=True, data_only=True)
                    log("✓ Normalized path access successful")
                except Exception as e3:
                    log(f"Normalized path access failed: {e3}")
                    log("All access methods failed - file not accessible")
                    return {}
        
        # Шукаємо вкладку "Tehnoloogiad"
        if "Tehnoloogiad" not in wb.sheetnames:
            log(f"Warning: Sheet 'Tehnoloogiad' not found. Available sheets: {wb.sheetnames}")
            wb.close()
            return {}
        
        ws = wb["Tehnoloogiad"]
        
        target_times = {}
        
        # Маппінг колонок для кожної операції: (станок_col, час_col)
        # K=11, N=14, P=16, S=19, U=21, X=24, Z=26, AC=29, AE=31, AH=34
        op_columns = {
            1: (11, 14),   # K, N
            2: (16, 19),   # P, S
            3: (21, 24),   # U, X
            4: (26, 29),   # Z, AC
            5: (31, 34),   # AE, AH
        }
        
        rows_processed = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Колонка A (індекс 0) - назва програми
            program_cell = row[0]
            if not program_cell or not program_cell.value:
                continue
            
            program_name = str(program_cell.value).strip()
            if not program_name:
                continue
            
            rows_processed += 1
            
            # Для кожної операції перевіряємо станок та час
            for op_num, (machine_col, time_col) in op_columns.items():
                machine_cell = row[machine_col - 1]  # -1 бо індексація з 0
                time_cell = row[time_col - 1]
                
                if machine_cell and machine_cell.value and time_cell and time_cell.value:
                    machine = str(machine_cell.value).strip()
                    try:
                        time_val = float(time_cell.value)
                        # Зберігаємо як: (program, operation, machine) -> time
                        key = (program_name, op_num, machine)
                        target_times[key] = time_val
                    except (ValueError, TypeError):
                        continue
        
        wb.close()
        log(f"✓ Loaded {len(target_times)} target times from {rows_processed} rows")
        return target_times
        
    except ImportError:
        log("Warning: openpyxl not installed. Install with: pip install openpyxl --break-system-packages")
        return {}
    except Exception as e:
        log(f"Error loading target times: {e}")
        import traceback
        log(traceback.format_exc())
        return {}

def get_actual_cycle_time_from_mr(machine_name, program_name, mr_data):
    """Бере фактичний час циклу з MachiningResult (RunStateTime / Counter)
    
    Це НАЙТОЧНІШИЙ метод - дані прямо з верстата.
    Береться МЕДІАНА з усіх записів для програми (щоб відсіяти паузи).
    
    Args:
        machine_name: повна назва машини (напр. "M1_M560R-V-e_0712-100198")
        program_name: назва програми (напр. "WF901-907-1.MIN")
        mr_data: список записів з machining_results.csv
    
    Returns:
        float: фактичний час одного циклу в хвилинах або None
    """
    if not mr_data:
        log(f"  get_actual_cycle_time_from_mr: mr_data is empty!")
        return None
    
    log(f"  get_actual_cycle_time_from_mr: machine={machine_name}, program={program_name}, mr_data records={len(mr_data)}")
    
    # Нормалізуємо назву програми для порівняння
    prog_base, prog_op = parse_program_name(program_name)
    
    # Збираємо ВСІ записи для цієї програми
    cycle_times = []
    
    for mr in mr_data:
        mr_machine = mr.get("MachineName", "")
        mr_prog = mr.get("ProgramFileName", "")
        
        if mr_machine != machine_name:
            continue
        
        # Порівнюємо програми
        mr_prog_base, mr_prog_op = parse_program_name(mr_prog)
        if mr_prog_base != prog_base or mr_prog_op != prog_op:
            continue
        
        # Знайшли запис для цієї програми
        run_time = int(mr.get("RunStateTime", 0))  # секунди
        counter = int(mr.get("Counter", 1))
        
        if run_time > 0 and counter > 0:
            cycle_time = run_time / counter / 60  # хвилини
            cycle_times.append(cycle_time)
            log(f"  • MR record: RunStateTime={run_time}s, Counter={counter} → {round(cycle_time, 2)} min/cycle")
    
    if not cycle_times:
        log(f"  ✗ No MR records found for {machine_name} / {program_name}")
        return None
    
    # Беремо МЕДІАНУ (щоб відсіяти викиди з паузами)
    sorted_times = sorted(cycle_times)
    n = len(sorted_times)
    
    if n == 1:
        result = sorted_times[0]
    elif n % 2 == 0:
        result = (sorted_times[n//2 - 1] + sorted_times[n//2]) / 2
    else:
        result = sorted_times[n//2]
    
    result = round(result, 2)
    log(f"  ✓ Median from {n} records = {result} min")
    
    return result


def filter_completed_cycles(cycles_list):
    """Фільтрує тільки 'хороші' завершені цикли для розрахунку
    
    Відсіює:
    - Перервані цикли (без end)
    - Розділені scheduled цикли (split_from_schedule)
    - Занадто короткі (<0.5 хв = 30 сек)
    
    Args:
        cycles_list: список циклів
    
    Returns:
        list: тривалості тільки 'хороших' циклів
    """
    good_durations = []
    
    for c in cycles_list:
        # Пропускаємо ongoing цикли
        if c.get("end") is None:
            continue
        
        # Пропускаємо розділені scheduled цикли
        if c.get("split_from_schedule"):
            continue
        
        # Пропускаємо занадто короткі (ймовірно перервані)
        duration = c.get("duration", 0)
        if duration < 0.5:  # менше 30 секунд
            continue
        
        good_durations.append(duration)
    
    return good_durations


def calculate_real_cycle_time(durations):
    """Визначає реальну довжину циклу методом медіани та щільного кластеру
    
    Args:
        durations: список тривалостей циклів (хвилини)
    
    Returns:
        float: реальна довжина циклу або None якщо недостатньо даних
    """
    if not durations or len(durations) < 3:
        return None
    
    # 1. Сортуємо
    sorted_durations = sorted(durations)
    n = len(sorted_durations)
    
    # 2. Знаходимо медіану всього масиву
    if n % 2 == 0:
        median = (sorted_durations[n//2 - 1] + sorted_durations[n//2]) / 2
    else:
        median = sorted_durations[n//2]
    
    # 3. Вікно допуску = медіана × 0.30
    window = median * 0.30
    
    # 4. Для кожного елементу рахуємо кількість сусідів
    neighbor_counts = []
    for i, value in enumerate(sorted_durations):
        count = sum(1 for d in sorted_durations if value - window <= d <= value + window)
        neighbor_counts.append((i, value, count))
    
    # 5. Знаходимо елемент(и) з максимальною кількістю сусідів
    max_neighbors = max(nc[2] for nc in neighbor_counts)
    centers = [nc for nc in neighbor_counts if nc[2] == max_neighbors]
    
    # Якщо кілька центрів - беремо той що ближче до середини списку
    if len(centers) > 1:
        middle_idx = n / 2
        center = min(centers, key=lambda nc: abs(nc[0] - middle_idx))
    else:
        center = centers[0]
    
    center_value = center[1]
    
    # 6. Зібрати всі елементи в діапазоні ±вікно навколо центру
    cluster = [d for d in sorted_durations if center_value - window <= d <= center_value + window]
    
    # 7. Медіана кластеру
    cluster_sorted = sorted(cluster)
    cluster_n = len(cluster_sorted)
    
    if cluster_n == 0:
        return None
    elif cluster_n % 2 == 0:
        result = (cluster_sorted[cluster_n//2 - 1] + cluster_sorted[cluster_n//2]) / 2
    else:
        result = cluster_sorted[cluster_n//2]
    
    return round(result, 2)


def calculate_cycle_time_smart(machine_name, program_name, cycles_list, mr_data):
    """РОЗУМНИЙ розрахунок часу циклу з пріоритетами
    
    Пріоритет 1: RunStateTime з MachiningResult (найточніше)
    Пріоритет 2: Кластеризація відфільтрованих циклів
    Пріоритет 3: None (недостатньо даних)
    
    Args:
        machine_name: назва машини
        program_name: назва програми
        cycles_list: список циклів
        mr_data: дані з MachiningResult
    
    Returns:
        float: розрахований час циклу або None
    """
    # Спроба 1: Взяти з MachiningResult
    actual_time = get_actual_cycle_time_from_mr(machine_name, program_name, mr_data)
    
    if actual_time is not None:
        return actual_time  # ✅ Найточніше!
    
    # Спроба 2: Фільтруємо та використовуємо кластеризацію
    good_durations = filter_completed_cycles(cycles_list)
    
    if len(good_durations) >= 3:
        return calculate_real_cycle_time(good_durations)  # ⚠️ Fallback
    
    # Спроба 3: Якщо і після фільтрації мало даних - беремо всі
    all_durations = [c["duration"] for c in cycles_list if c["duration"] > 0]
    
    if len(all_durations) >= 3:
        return calculate_real_cycle_time(all_durations)  # ⚠️⚠️ Last resort
    
    return None  # ❌ Недостатньо даних


def calculate_real_cycle_time_OLD(durations):
    """Визначає реальну довжину циклу методом медіани та щільного кластеру
    
    Args:
        durations: список тривалостей циклів (хвилини)
    
    Returns:
        float: реальна довжина циклу або None якщо недостатньо даних
    """
    if not durations or len(durations) < 3:
        return None
    
    # 1. Сортуємо
    sorted_durations = sorted(durations)
    n = len(sorted_durations)
    
    # 2. Знаходимо медіану всього масиву
    if n % 2 == 0:
        median = (sorted_durations[n//2 - 1] + sorted_durations[n//2]) / 2
    else:
        median = sorted_durations[n//2]
    
    # 3. Вікно допуску = медіана × 0.30
    window = median * 0.30
    
    # 4. Для кожного елементу рахуємо кількість сусідів
    neighbor_counts = []
    for i, value in enumerate(sorted_durations):
        count = sum(1 for d in sorted_durations if value - window <= d <= value + window)
        neighbor_counts.append((i, value, count))
    
    # 5. Знаходимо елемент(и) з максимальною кількістю сусідів
    max_neighbors = max(nc[2] for nc in neighbor_counts)
    centers = [nc for nc in neighbor_counts if nc[2] == max_neighbors]
    
    # Якщо кілька центрів - беремо той що ближче до середини списку
    if len(centers) > 1:
        middle_idx = n / 2
        center = min(centers, key=lambda nc: abs(nc[0] - middle_idx))
    else:
        center = centers[0]
    
    center_value = center[1]
    
    # 6. Зібрати всі елементи в діапазоні ±вікно навколо центру
    cluster = [d for d in sorted_durations if center_value - window <= d <= center_value + window]
    
    # 7. Медіана кластеру
    cluster_sorted = sorted(cluster)
    cluster_n = len(cluster_sorted)
    
    if cluster_n == 0:
        return None
    elif cluster_n % 2 == 0:
        result = (cluster_sorted[cluster_n//2 - 1] + cluster_sorted[cluster_n//2]) / 2
    else:
        result = cluster_sorted[cluster_n//2]
    
    return round(result, 1)

# PART 1 — DOWNLOAD
# =============================================================================

# =============================================================================
# PART 1 — DOWNLOAD (інтегрована функція download_both_files)
# =============================================================================

def download_both_files():
    """Завантажує operation_history.csv та machining_results.csv за один запуск браузера"""
    log("============================================================")
    log("DOWNLOADING BOTH FILES FROM CONNECT PLAN")
    log("============================================================")
    
    # Крок 1: Видаляємо всі CSV файли
    log("── Step 1: Clearing all CSV files ──")
    for filename in os.listdir(DOWNLOAD_DIR):
        if filename.lower().endswith(".csv"):
            try:
                os.remove(os.path.join(DOWNLOAD_DIR, filename))
                log(f"Deleted: {filename}")
            except Exception as exc:
                log(f"Could not delete {filename}: {exc}")
    
    # Налаштування Chrome
    options = Options()
    options.add_experimental_option("prefs", {
        "download.default_directory":   DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade":   True,
        "safebrowsing.enabled":         True,
    })
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    
    try:
        # Крок 2: Відкриваємо сторінку
        log("── Step 2: Opening page in Chrome ──")
        driver.get(URL)
        wait = WebDriverWait(driver, WAIT_TIME)
        time.sleep(3)
        
        # Крок 3: Обираємо всі машини
        log("── Step 3: Selecting all machines ──")
        checkbox = wait.until(EC.element_to_be_clickable((By.ID, "all_machines_check")))
        if not checkbox.is_selected():
            checkbox.click()
            time.sleep(3)
            log("✓ All machines selected")
        else:
            log("✓ All machines already selected")
        
        # Крок 4: Тиснемо Download (operation_history)
        log("── Step 4: Downloading operation_history.csv ──")
        btn = wait.until(EC.element_to_be_clickable((By.ID, "btn_Download")))
        click_time_1 = time.time()
        btn.click()
        log("✓ Download button clicked — waiting for first file...")
        
        # Чекаємо перший файл
        deadline = time.time() + WAIT_TIME
        downloaded_1 = None
        while time.time() < deadline:
            for f in os.listdir(DOWNLOAD_DIR):
                if f.lower().endswith(".csv"):
                    fp = os.path.join(DOWNLOAD_DIR, f)
                    if os.path.getmtime(fp) >= click_time_1:
                        downloaded_1 = fp
                        break
            if downloaded_1:
                break
            time.sleep(2)
            print(".", end="", flush=True)
        
        print()
        
        if not downloaded_1:
            log("✗ Error: First file did not download")
            return False
        
        log(f"✓ First file downloaded: {os.path.basename(downloaded_1)}")
        
        # Крок 5: Вибираємо MACHINING RESULT з dropdown
        log("── Step 5: Selecting MACHINING RESULT from dropdown ──")
        time.sleep(2)  # Даємо час завантажитись попередньому файлу
        
        try:
            dropdown = wait.until(EC.presence_of_element_located((By.ID, "ddl_SelectTable")))
            select = Select(dropdown)
            select.select_by_value("MachningResult")  # З опечаткою як в HTML!
            log("✓ Selected: MACHINING RESULT")
            time.sleep(3)  # Чекаємо на __doPostBack
        except Exception as e:
            log(f"✗ Could not select MACHINING RESULT: {e}")
            return False
        
        # Крок 6: Тиснемо Download (MachiningResult)
        log("── Step 6: Downloading machining_results.csv ──")
        btn = wait.until(EC.element_to_be_clickable((By.ID, "btn_Download")))
        click_time_2 = time.time()
        btn.click()
        log("✓ Download button clicked — waiting for second file...")
        
        # Чекаємо другий файл
        deadline = time.time() + WAIT_TIME
        downloaded_2 = None
        while time.time() < deadline:
            for f in os.listdir(DOWNLOAD_DIR):
                if f.lower().endswith(".csv"):
                    fp = os.path.join(DOWNLOAD_DIR, f)
                    # Шукаємо файл новіший за другий клік і не той що вже downloaded_1
                    if os.path.getmtime(fp) >= click_time_2 and fp != downloaded_1:
                        downloaded_2 = fp
                        break
            if downloaded_2:
                break
            time.sleep(2)
            print(".", end="", flush=True)
        
        print()
        
        if not downloaded_2:
            log("✗ Error: Second file did not download")
            return False
        
        log(f"✓ Second file downloaded: {os.path.basename(downloaded_2)}")
        
        # Крок 7: Перейменовуємо файли
        log("── Step 7: Renaming files ──")
        
        # Перейменовуємо перший файл в operation_history.csv
        if os.path.exists(CSV_FILE):
            os.remove(CSV_FILE)
        os.rename(downloaded_1, CSV_FILE)
        log(f"✓ Renamed to: operation_history.csv")
        
        # Показуємо інфо про перший файл
        size1 = os.path.getsize(CSV_FILE)
        with open(CSV_FILE, 'r', encoding='utf-8') as f:
            lines1 = sum(1 for _ in f)
        log(f"  Size: {size1:,} bytes, Rows: {lines1}")
        
        # Перейменовуємо другий файл в machining_results.csv
        if os.path.exists(MACHINING_RESULT):
            os.remove(MACHINING_RESULT)
        os.rename(downloaded_2, MACHINING_RESULT)
        log(f"✓ Renamed to: machining_results.csv")
        
        # Показуємо інфо про другий файл
        size2 = os.path.getsize(MACHINING_RESULT)
        with open(MACHINING_RESULT, 'r', encoding='utf-8') as f:
            lines2 = sum(1 for _ in f)
        log(f"  Size: {size2:,} bytes, Rows: {lines2}")
        
        log("============================================================")
        log("DOWNLOAD COMPLETE")
        log("============================================================")
        return True
        
    except TimeoutException:
        log("✗ Error: element not found on page")
        return False
    except WebDriverException as exc:
        log(f"✗ WebDriver error: {exc}")
        return False
    except Exception as exc:
        log(f"✗ Unexpected error: {exc}")
        import traceback
        log(traceback.format_exc())
        return False
    finally:
        log("Closing browser...")
        time.sleep(2)
        driver.quit()

# =============================================================================
# PART 2 — ANALYSIS
# =============================================================================

# ── Telegram ──────────────────────────────────────────────────────────────────
def send_telegram(message: str):
    try:
        url  = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        data = urllib.parse.urlencode({
            "chat_id":    TELEGRAM_CHAT_ID,
            "text":       message,
            "parse_mode": "HTML",
        }).encode()
        urllib.request.urlopen(urllib.request.Request(url, data=data), timeout=10)
        log("Telegram alert sent")
    except Exception as e:
        log(f"Telegram error: {e}")

# ── SQLite ────────────────────────────────────────────────────────────────────
def init_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS daily_summary (
            date TEXT, machine TEXT, run_min INTEGER, down_min INTEGER,
            total_min INTEGER, cycles INTEGER, avg_cycle REAL, efficiency REAL,
            PRIMARY KEY (date, machine)
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS downtime_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, machine TEXT, start_time TEXT, end_time TEXT,
            duration INTEGER, reason TEXT
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cycle_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, machine TEXT, program TEXT, start_time TEXT, end_time TEXT,
            duration INTEGER
        )""")

    conn.commit()
    return conn


def save_to_db(conn, date_str, cycles, downtimes):
    for mname in cycles:
        c_list    = cycles[mname]
        d_data    = downtimes[mname]
        run_min   = d_data["total_run"]
        down_min  = d_data["total_down"]
        total_min = d_data["total_min"]
        eff       = round(run_min / total_min * 100, 1) if total_min else 0
        avg_cycle = round(sum(c["duration"] for c in c_list) / len(c_list), 1) if c_list else 0
        conn.execute("""
            INSERT OR REPLACE INTO daily_summary
            (date,machine,run_min,down_min,total_min,cycles,avg_cycle,efficiency)
            VALUES (?,?,?,?,?,?,?,?)
        """, (date_str, mname, run_min, down_min, total_min, len(c_list), avg_cycle, eff))
        
        # Зберігаємо окремі цикли
        for c in c_list:
            conn.execute("""
                INSERT OR IGNORE INTO cycle_events
                (date,machine,program,start_time,end_time,duration)
                VALUES (?,?,?,?,?,?)
            """, (date_str, mname, c.get("program", "—"),
                  c["start"].strftime("%H:%M") if c.get("start") else "—",
                  c["end"].strftime("%H:%M") if c.get("end") else "—",
                  c["duration"]))
        
        # Зберігаємо простої
        for d in d_data["downtimes"]:
            conn.execute("""
                INSERT OR IGNORE INTO downtime_events
                (date,machine,start_time,end_time,duration,reason)
                VALUES (?,?,?,?,?,?)
            """, (date_str, mname,
                  d["start"].strftime("%H:%M"),
                  d["end"].strftime("%H:%M") if d.get("end") else "ongoing",
                  d["duration"], d["reason"]))

    conn.commit()

def load_history(conn, machine: str, days: int = 7) -> list:
    cur = conn.execute("""
        SELECT date, efficiency, run_min, down_min, cycles, avg_cycle
        FROM daily_summary WHERE machine=? ORDER BY date DESC LIMIT ?
    """, (machine, days))
    return list(reversed(cur.fetchall()))

def get_recent_cycles(conn, machine: str, program: str, current_cycles: list, target_count: int = 20) -> list:
    """Отримує цикли для програми: останні 3 робочі дні станку + з історії якщо потрібно
    
    Args:
        conn: з'єднання з БД
        machine: назва станку
        program: назва програми
        current_cycles: поточні цикли за сьогодні (список dict з duration)
        target_count: цільова кількість циклів (за замовчуванням 20)
    
    Returns:
        list: список тривалостей циклів (int)
    """
    from datetime import datetime
    
    # Збираємо тривалості з поточних циклів
    durations = [c["duration"] for c in current_cycles]
    
    # Знаходимо останні 3 РОБОЧІ ДНІ СТАНКУ (дні коли станок працював, будь-які програми)
    today = datetime.now().strftime("%Y-%m-%d")
    cur = conn.execute("""
        SELECT DISTINCT date 
        FROM cycle_events 
        WHERE machine = ? AND date != ?
        ORDER BY date DESC 
        LIMIT 3
    """, (machine, today))
    
    last_3_working_days = [row[0] for row in cur.fetchall()]
    
    if not last_3_working_days:
        # Немає робочих днів - повертаємо поточні або добираємо з історії
        if len(durations) >= target_count:
            return durations
        # Добираємо з усієї історії
        needed = target_count - len(durations)
        cur = conn.execute("""
            SELECT duration 
            FROM cycle_events 
            WHERE machine = ? AND program = ? AND date != ?
            ORDER BY date DESC, start_time DESC
            LIMIT ?
        """, (machine, program, today, needed))
        historical = [row[0] for row in cur.fetchall()]
        return durations + historical
    
    # Завантажуємо цикли ЦІЄЇ ПРОГРАМИ з останніх 3 робочих днів
    placeholders = ','.join('?' * len(last_3_working_days))
    cur = conn.execute(f"""
        SELECT duration 
        FROM cycle_events 
        WHERE machine = ? AND program = ? AND date IN ({placeholders})
        ORDER BY date DESC, start_time DESC
    """, (machine, program, *last_3_working_days))
    
    cycles_from_3_days = [row[0] for row in cur.fetchall()]
    
    # Об'єднуємо: поточні + з останніх 3 днів
    all_cycles = durations + cycles_from_3_days
    
    # ВАЖЛИВО: Якщо за останні 3 дні >= 20 циклів - беремо ВСІ з цього періоду
    if len(all_cycles) >= target_count:
        return all_cycles
    
    # Якщо менше 20 - добираємо з історії (старіші дні)
    used_days = [today] + last_3_working_days
    used_days_placeholders = ','.join('?' * len(used_days))
    
    needed = target_count - len(all_cycles)
    cur = conn.execute(f"""
        SELECT duration 
        FROM cycle_events 
        WHERE machine = ? AND program = ? AND date NOT IN ({used_days_placeholders})
        ORDER BY date DESC, start_time DESC
        LIMIT ?
    """, (machine, program, *used_days, needed))
    
    historical = [row[0] for row in cur.fetchall()]
    all_cycles.extend(historical)
    
    return all_cycles

# ── Data processing ───────────────────────────────────────────────────────────
def load_csv() -> list[dict]:
    with open(CSV_FILE, encoding="utf-8") as f:
        return list(csv.DictReader(f))

def filter_last_hours(rows, hours):
    parse   = lambda s: datetime.strptime(s, "%Y.%m.%d %H:%M:%S")
    last_ts = max(parse(r["Date"]) for r in rows)
    # Для 24 годин — починаємо з 00:00 того ж дня
    if hours >= 24:
        cutoff = last_ts.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        cutoff = last_ts - timedelta(hours=hours)
    return [r for r in rows if parse(r["Date"]) >= cutoff], cutoff, last_ts

def analyze_cycles(rows):
    parse    = lambda s: datetime.strptime(s, "%Y.%m.%d %H:%M:%S")
    machines = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)
    result = {}
    for mname, mrows in machines.items():
        mrows.sort(key=lambda r: r["Date"])
        cycles, prev_run, prev_prog_parsed, cycle_start, cycle_prog = [], None, None, None, ""
        for r in mrows:
            ts, run, prog = parse(r["Date"]), r["RunState"], r["ProgramFileName"]
            prog_parsed = parse_program_name(prog)  # (base, operation)
            
            # Фіксуємо старт циклу коли:
            # 1. RunState змінився з 0→1 АБО
            # 2. Програма змінилась при RunState=1
            if run == "1":
                if prev_run in (None, "0"):
                    # Старт нового циклу
                    cycle_start, cycle_prog = ts, prog
                elif prev_run == "1" and prog_parsed != prev_prog_parsed and cycle_start:
                    # Програма змінилась - закриваємо попередній цикл
                    cycles.append({"start": cycle_start, "end": ts, "program": cycle_prog,
                                    "duration": round((ts - cycle_start).total_seconds() / 60, 2)})
                    # Стартуємо новий цикл з новою програмою
                    cycle_start, cycle_prog = ts, prog
            
            # Фіксуємо кінець циклу коли RunState змінився з 1→0
            elif prev_run == "1" and run == "0" and cycle_start:
                cycles.append({"start": cycle_start, "end": ts, "program": cycle_prog,
                                "duration": round((ts - cycle_start).total_seconds() / 60, 2)})
                cycle_start = None
            
            prev_run = run
            prev_prog_parsed = prog_parsed
            
        if cycle_start:
            last_ts = parse(mrows[-1]["Date"])
            cycles.append({"start": cycle_start, "end": None, "program": cycle_prog,
                           "duration": round((last_ts - cycle_start).total_seconds() / 60, 2),
                           "ongoing": True})
        result[mname] = cycles
    return result

def split_schedule_cycles(cycles_dict, machining_result_file):
    """Розділяє цикли schedule program на окремі цикли на основі Counter з MachiningResult
    
    Args:
        cycles_dict: словник циклів {machine: [cycles]}
        machining_result_file: шлях до MachiningResult CSV файлу
    
    Returns:
        tuple: (cycles_dict, mr_data) - оновлений словник циклів та дані MachiningResult
    """
    import csv
    from datetime import timedelta
    
    # Якщо файл не існує - повертаємо як є
    if not os.path.exists(machining_result_file):
        log(f"MachiningResult file not found: {machining_result_file}")
        return cycles_dict, []
    
    try:
        # Завантажуємо MachiningResult
        with open(machining_result_file, encoding="utf-8") as f:
            mr_data = list(csv.DictReader(f))
        
        log(f"Loaded {len(mr_data)} records from MachiningResult")
        
        # Для кожного станку
        for mname, cycles in cycles_dict.items():
            new_cycles = []
            
            for cycle in cycles:
                prog = cycle.get("program", "")
                cycle_start = cycle["start"]
                cycle_end = cycle.get("end")
                cycle_duration = cycle["duration"]
                is_ongoing = cycle.get("ongoing", False)
                
                # Ongoing цикли НЕ розділяємо - залишаємо як є
                if is_ongoing:
                    new_cycles.append(cycle)
                    continue
                
                # Шукаємо відповідний запис у MachiningResult
                counter = 1  # За замовчуванням 1 цикл (обнуляємо для кожного циклу!)
                matched_mr = None  # Зберігаємо знайдений запис
                
                # Розбираємо назву програми: базова назва + операція
                prog_base, prog_op = parse_program_name(prog)
                
                for mr in mr_data:
                    mr_machine = mr.get("MachineName", "")
                    mr_prog = mr.get("ProgramFileName", "")
                    mr_prog_base, mr_prog_op = parse_program_name(mr_prog)
                    mr_date_str = mr.get("Date", "")
                    
                    # Перевіряємо чи це наш станок, програма І операція
                    if mr_machine == mname and mr_prog_base == prog_base and mr_prog_op == prog_op:
                        try:
                            mr_date = datetime.strptime(mr_date_str, "%Y.%m.%d %H:%M:%S")
                            # Якщо час MachiningResult близький до кінця циклу (±5 хв)
                            if cycle_end and abs((mr_date - cycle_end).total_seconds()) < 300:
                                counter = int(mr.get("Counter", 1))
                                matched_mr = mr  # Зберігаємо запис
                                if counter > 1:
                                    log(f"{mname} {prog}: Found Counter={counter}, splitting cycle")
                                break
                        except:
                            continue
                
                # Якщо Counter > 1 - розділяємо цикл
                if counter > 1 and cycle_duration > 0 and matched_mr:
                    # Беремо ФАКТИЧНИЙ час з MachiningResult
                    run_state_time = int(matched_mr.get("RunStateTime", 0))  # в секундах
                    
                    if run_state_time > 0:
                        # Фактична тривалість одного циклу
                        sub_duration = round(run_state_time / counter / 60, 2)  # секунди → хвилини
                        log(f"{mname} {prog}: Actual cycle time = {sub_duration} min (from RunStateTime={run_state_time}s / {counter})")
                    else:
                        # Fallback: ділимо загальний час циклу
                        sub_duration = round(cycle_duration / counter, 2)
                        log(f"{mname} {prog}: No RunStateTime, using calculated {sub_duration} min")
                    
                    for i in range(counter):
                        # Для розділених циклів start/end = None (буде показано як "—")
                        # Тільки duration має значення
                        new_cycles.append({
                            "start": None,  # Прочерк
                            "end": None,    # Прочерк
                            "program": prog,
                            "duration": sub_duration,
                            "ongoing": False,  # Розділені цикли не ongoing
                            "split_from_schedule": True  # Позначка що це розділений цикл
                        })
                else:
                    # Counter = 1 або не знайдено - залишаємо як є
                    new_cycles.append(cycle)
            
            cycles_dict[mname] = new_cycles
        
        return cycles_dict, mr_data  # Повертаємо і cycles і mr_data
        
    except Exception as e:
        log(f"Error splitting schedule cycles: {e}")
        import traceback
        log(traceback.format_exc())
        return cycles_dict, []  # При помилці повертаємо порожній mr_data

def analyze_downtime(rows):
    parse    = lambda s: datetime.strptime(s, "%Y.%m.%d %H:%M:%S")
    machines = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)
    result = {}
    for mname, mrows in machines.items():
        mrows.sort(key=lambda r: r["Date"])
        downtimes, prev_run, dt_start, dt_reason = [], None, None, ""
        
        # Для розрахунку ефективності виключаємо нічний період
        filtered_rows = []
        for r in mrows:
            ts = parse(r["Date"])
            hour = ts.hour + ts.minute / 60.0
            # Виключаємо 01:00 - 06:30 (з 1.0 до 6.5 години)
            if not (1.0 <= hour < 6.5):
                filtered_rows.append(r)
        
        for r in mrows:
            ts, run = parse(r["Date"]), r["RunState"]
            if run == "0":
                if   r["AlarmState"]       == "1": reason = "Alarm: " + (r["AlarmMessage"] or r["AlarmNo"] or "—")
                elif r["PowerOn"]          == "0": reason = "Power off"
                elif r["SetUp"]            == "1": reason = "Setup"
                elif r["Maintenance"]      == "1": reason = "Maintenance"
                elif r["NoOperator"]       == "1": reason = "No operator"
                elif r["Wait"]             == "1": reason = "Waiting"
                elif r["FeedHoldState"]    == "1": reason = "Feed Hold"
                elif r["ProgramStopState"] == "1": reason = "Program Stop"
                else:                              reason = "Idle"
            else:
                reason = ""
            if prev_run in (None, "1") and run == "0":
                dt_start, dt_reason = ts, reason
            elif prev_run == "0" and run == "1" and dt_start:
                dur = round((ts - dt_start).total_seconds() / 60, 2)
                if dur > 0:
                    downtimes.append({"start": dt_start, "end": ts,
                                      "duration": dur, "reason": dt_reason})
                dt_start = None
            prev_run = run
        if dt_start:
            last_ts = parse(mrows[-1]["Date"])
            dur = round((last_ts - dt_start).total_seconds() / 60, 2)
            if dur > 0:
                downtimes.append({"start": dt_start, "end": None, "duration": dur,
                                  "reason": dt_reason, "ongoing": True})
        
        # Рахуємо ефективність БЕЗ нічного періоду
        result[mname] = {
            "downtimes":  downtimes,
            "total_run":  sum(1 for r in filtered_rows if r["RunState"] == "1"),
            "total_down": sum(1 for r in filtered_rows if r["RunState"] == "0"),
            "total_min":  len(filtered_rows),
        }
    return result

def build_timeline_data(rows, period_from, period_to):
    parse     = lambda s: datetime.strptime(s, "%Y.%m.%d %H:%M:%S")
    machines  = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)
    total_sec = max((period_to - period_from).total_seconds(), 1)
    result    = {}

    for mname, mrows in machines.items():
        mrows.sort(key=lambda r: r["Date"])
        segments  = []
        seg_start = period_from
        seg_state = None
        seg_label = ""
        seg_idx   = 0

        def _get_label(r):
            run = r["RunState"]
            if run == "1":
                return r["ProgramFileName"] or "Running"
            if   r["AlarmState"]       == "1": return "Alarm: " + (r["AlarmMessage"] or r["AlarmNo"] or "—")
            elif r["PowerOn"]          == "0": return "Power off"
            elif r["SetUp"]            == "1": return "Setup"
            elif r["Maintenance"]      == "1": return "Maintenance"
            elif r["NoOperator"]       == "1": return "No operator"
            elif r["Wait"]             == "1": return "Waiting"
            elif r["FeedHoldState"]    == "1": return "Feed Hold"
            elif r["ProgramStopState"] == "1": return "Program Stop"
            return "Idle"

        for r in mrows:
            ts, run = parse(r["Date"]), r["RunState"]
            lbl = _get_label(r)
            if seg_state is None:
                seg_state, seg_start, seg_label = run, ts, lbl
            elif run != seg_state or (run == "1" and lbl != seg_label):
                # нова програма або зміна стану — закриваємо сегмент
                x = (seg_start - period_from).total_seconds() / total_sec * 100
                w = (ts - seg_start).total_seconds() / total_sec * 100
                if w > 0.05:
                    segments.append({
                        "x": x, "w": w, "state": seg_state,
                        "label": seg_label,
                        "start": seg_start.strftime("%H:%M"),
                        "end":   ts.strftime("%H:%M"),
                        "id":    f"{mname.split('_')[0]}_{seg_idx}",
                    })
                    seg_idx += 1
                seg_start, seg_state, seg_label = ts, run, lbl

        if seg_state is not None:
            x = (seg_start - period_from).total_seconds() / total_sec * 100
            w = (period_to - seg_start).total_seconds() / total_sec * 100
            if w > 0.05:
                segments.append({
                    "x": x, "w": w, "state": seg_state,
                    "label": seg_label,
                    "start": seg_start.strftime("%H:%M"),
                    "end":   period_to.strftime("%H:%M"),
                    "id":    f"{mname.split('_')[0]}_{seg_idx}",
                })
        result[mname] = segments
    return result

# ── GitHub Pages publish ──────────────────────────────────────────────────────
def publish_to_github(html: str) -> bool:
    """Push index.html to GitHub Pages via API — no git install required."""
    import base64
    import traceback
    try:
        api     = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/index.html"
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Content-Type":  "application/json",
            "Accept":        "application/vnd.github+json",
        }
        log(f"GitHub API URL: {api}")
        log(f"GitHub User: {GITHUB_USER}")
        log(f"GitHub Repo: {GITHUB_REPO}")
        
        # Отримуємо SHA якщо файл вже існує
        sha = None
        try:
            log("Checking if index.html exists...")
            req = urllib.request.Request(api, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as r:
                data = json.loads(r.read().decode())
                sha = data["sha"]
                log(f"File exists, SHA: {sha[:8]}...")
        except urllib.error.HTTPError as e:
            if e.code == 404:
                log("File doesn't exist, will create new")
            else:
                log(f"HTTP Error checking file: {e.code} {e.reason}")
                raise
        
        # Пушимо файл
        payload = {
            "message": f"update {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "content": base64.b64encode(html.encode("utf-8")).decode(),
            "branch": "main"  # явно вказуємо гілку
        }
        if sha:
            payload["sha"] = sha
        
        log(f"Uploading HTML ({len(html)} bytes)...")
        req = urllib.request.Request(
            api, data=json.dumps(payload).encode(),
            headers=headers, method="PUT"
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            response = json.loads(r.read().decode())
            log(f"Upload successful!")
            log(f"Commit SHA: {response['commit']['sha'][:8]}...")
        
        log(f"✓ Published: {GITHUB_URL}")
        return True
    except urllib.error.HTTPError as e:
        log(f"✗ GitHub HTTP Error: {e.code} {e.reason}")
        try:
            error_body = e.read().decode()
            log(f"Error details: {error_body}")
        except:
            pass
        return False
    except Exception as e:
        log(f"✗ GitHub publish error: {type(e).__name__}: {e}")
        import io
        s = io.StringIO()
        traceback.print_exc(file=s)
        log(s.getvalue())
        return False

def check_and_alert(downtimes, period_to, cycles, excel_targets, mr_data):
    """Перевіряє простої та відправляє Telegram алерти
    
    Умови відправлення алерту:
    1. Є новий невідрапортований простій ≥45 хв, АБО
    2. Є старий ongoing простій що збільшився на +45 хв, АБО
    3. Є різниця між Calculated та Target >10%
    
    Додаткові правила:
    - Не відправляємо з 20:00 до 08:00
    - Закінчені простої не повторюємо
    - Target алерти відправляємо раз на добу
    - Через 24 години скидаємо список
    """
    from datetime import datetime, timedelta
    import json
    
    log("── Step 3.6: Checking alerts ──")
    
    current_hour = datetime.now().hour
    current_time = datetime.now()
    
    log(f"Current hour: {current_hour}, downtimes: {len(downtimes)} machines, cycles: {len(cycles)} machines")
    
    # 1. Перевіряємо тихі години (20:00 - 08:00)
    if current_hour >= 20 or current_hour < 8:
        log("Silent hours (20:00-08:00) - no alerts sent")
        return
    
    # Файл для збереження вже відправлених алертів
    sent_alerts_file = os.path.join(DOWNLOAD_DIR, "sent_alerts.json")
    
    # Завантажуємо список відправлених алертів
    sent_alerts = {}
    reset_needed = False
    
    if os.path.exists(sent_alerts_file):
        try:
            with open(sent_alerts_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                last_reset = data.get("last_reset", "")
                
                if last_reset:
                    last_reset_dt = datetime.fromisoformat(last_reset)
                    hours_since_reset = (current_time - last_reset_dt).total_seconds() / 3600
                    if hours_since_reset > 24:
                        reset_needed = True
                    else:
                        sent_alerts = data.get("alerts", {})
                else:
                    reset_needed = True
        except:
            reset_needed = True
    else:
        reset_needed = True
    
    if reset_needed:
        sent_alerts = {}
    
    # Збираємо алерти про простої
    downtime_alerts = []
    
    log(f"Checking downtimes for {len(downtimes)} machines...")
    for mname, dd in downtimes.items():
        machine_downtimes = dd.get("downtimes", [])
        log(f"  {mname}: {len(machine_downtimes)} downtime events")
        
        for d in machine_downtimes:
            duration = d["duration"]
            is_ongoing = not d.get("end")
            log(f"    - Duration: {duration} min, Ongoing: {is_ongoing}, Threshold: {ALERT_THRESHOLD_MIN} min")
            
            if duration >= ALERT_THRESHOLD_MIN:
                # Унікальний ключ: машина + час початку (не змінюється для ongoing)
                alert_key = f"downtime_{mname}_{d['start'].strftime('%Y-%m-%d_%H:%M')}"
                already_sent = alert_key in sent_alerts
                
                log(f"    - Qualifies for alert! Key: {alert_key}, Already sent: {already_sent}")
                
                # Перевіряємо чи вже відправляли алерт для цього простою
                if already_sent:
                    # Якщо це ongoing простій - перевіряємо чи треба повторити
                    if is_ongoing:
                        prev_alert = sent_alerts[alert_key]
                        prev_duration = prev_alert.get("duration", 0)
                        additional_duration = duration - prev_duration
                        
                        log(f"    - Ongoing: prev_duration={prev_duration}, additional={additional_duration}")
                        
                        # Якщо протривав ще мінімум 45 хв - повторюємо алерт
                        if additional_duration >= 45:
                            log(f"    - ✓ Adding repeat alert (additional {additional_duration} min)")
                            downtime_alerts.append((mname, d, alert_key, True))  # True = repeat
                        else:
                            log(f"    - ✗ Not enough additional time ({additional_duration} < 45 min)")
                    else:
                        log(f"    - ✗ Downtime finished, not repeating")
                    # Якщо простій закінчився (є end) - НЕ повторюємо, пропускаємо
                else:
                    # Новий простій - відправляємо
                    log(f"    - ✓ Adding new alert")
                    downtime_alerts.append((mname, d, alert_key, False))  # False = new
            else:
                log(f"    - ✗ Below threshold ({duration} < {ALERT_THRESHOLD_MIN} min)")
    
    log(f"Found {len(downtime_alerts)} downtime alerts")
    
    # Збираємо алерти про перевищення Target >10%
    target_alerts = []
    machines_checked = set()  # Унікальні машини які перевірялись
    machines_with_issues = set()  # Машини з проблемами
    
    log(f"Checking cycle times for {len(cycles)} machines...")
    for mname, c_list in cycles.items():
        machine_short = mname.split("_")[0] if "_" in mname else mname
        
        # Групуємо по програмах
        by_prog = {}
        for c in c_list:
            prog_name = c["program"] or "—"
            if prog_name not in by_prog:
                by_prog[prog_name] = []
            by_prog[prog_name].append(c)
        
        for prog, prog_cycles in by_prog.items():
            # Розраховуємо Calculated РОЗУМНО (пріоритет: MachiningResult → фільтровані → всі)
            calc_target = calculate_cycle_time_smart(mname, prog, prog_cycles, mr_data)
            
            if calc_target is None:
                continue
            
            # Визначаємо операцію
            op_num = get_operation_number(prog)
            prog_normalized = normalize_program_name(prog)
            
            # Шукаємо Excel Target
            excel_target = None
            for (excel_prog, excel_op, excel_machine), time_val in excel_targets.items():
                excel_prog_normalized = normalize_program_name(excel_prog)
                excel_machine_normalized = normalize_program_name(excel_machine)
                machine_normalized = normalize_program_name(machine_short)
                
                if (prog_normalized == excel_prog_normalized and 
                    op_num == excel_op and 
                    machine_normalized == excel_machine_normalized):
                    excel_target = time_val
                    break
            
            # Якщо є Target
            if excel_target:
                machines_checked.add(machine_short)  # Додаємо до перевірених
                diff_pct = ((calc_target - excel_target) / excel_target) * 100
                log(f"    {prog}: Calculated={calc_target}, Target={excel_target}, Diff={round(diff_pct, 1)}%")
                
                if abs(diff_pct) > 10:
                    # Унікальний ключ для target алертів (для логування)
                    target_key = f"target_{mname}_{prog}"
                    
                    log(f"    ✓ Adding target alert (difference >10%)")
                    # ЗАВЖДИ додаємо - відправляємо кожен раз незалежно від історії
                    target_alerts.append((machine_short, prog, calc_target, excel_target, diff_pct, target_key))
                    machines_with_issues.add(machine_short)  # Додаємо до проблемних
                else:
                    log(f"    ✗ Difference ≤10%")
            else:
                log(f"    {prog}: No target found in Excel")
    
    log(f"Found {len(target_alerts)} target alerts")
    
    # Підраховуємо статистику
    total_machines = len(machines_checked)
    machines_with_issues_count = len(machines_with_issues)
    machines_ok_count = total_machines - machines_with_issues_count
    
    # Формуємо повідомлення
    if downtime_alerts or target_alerts:
        # Є проблеми
        lines = [
            f"⚠️ <b>Factory Alert</b>  {period_to.strftime('%H:%M')}",
            f"🔗 <a href=\"{GITHUB_URL}\">Open report</a>\n"
        ]
        
        # Summary
        status_parts = []
        if machines_ok_count > 0:
            status_parts.append(f"✅ {machines_ok_count} OK")
        if machines_with_issues_count > 0:
            status_parts.append(f"🔴 {machines_with_issues_count} need attention")
        
        if total_machines > 0:
            lines.append(f"📊 <b>Status:</b> {' | '.join(status_parts)}")
            lines.append("")  # Порожня лінія
    else:
        # Все ОК - відправляємо позитивне повідомлення
        lines = [
            f"✅ <b>All Systems Normal</b>  {period_to.strftime('%H:%M')}",
            f"🔗 <a href=\"{GITHUB_URL}\">Open report</a>\n",
            f"📊 All {total_machines} machines within target cycle times"
        ]
    
    log(f"Preparing to send: {len(downtime_alerts)} downtime alerts, {len(target_alerts)} target alerts")
    
    # Додаємо алерти про простої
    if downtime_alerts:
        for mname, d, alert_key, is_repeat in downtime_alerts:
            short = mname.split("_")[0]
            end_s = d["end"].strftime("%H:%M") if d.get("end") else "ongoing"
            
            # Позначка: 🔴🔴 для повторних, 🔴 для нових
            marker = "🔴🔴" if is_repeat else "🔴"
            
            lines.append(
                f"\n{marker} <b>{short}</b>  {d['start'].strftime('%H:%M')}–{end_s}"
                f"  <b>{d['duration']} min</b>\n   Reason: {d['reason']}"
            )
            
            # Оновлюємо інформацію про алерт
            sent_alerts[alert_key] = {
                "machine": mname,
                "start": d['start'].strftime('%Y-%m-%d %H:%M'),
                "duration": d['duration'],  # Поточна тривалість
                "last_alert": current_time.isoformat()
            }
    
    # Додаємо алерти про перевищення Target
    if target_alerts:
        lines.append("\n\n⚙️ <b>Cycle Time Alerts (>10% difference):</b>")
        for machine, prog, calc, target, diff_pct, target_key in target_alerts:
            abs_diff = abs(diff_pct)
            if diff_pct > 0:
                # Повільніше - червоний
                status = f'<b>{round(abs_diff, 1)}% slower</b> 🔴'
            else:
                # Швидше - зелений
                status = f'<b>{round(abs_diff, 1)}% faster</b> 🟢'
            
            lines.append(
                f"\n📊 <b>{machine}</b> - {prog}"
                f"\n   Calculated: {calc} min | Target: {target} min"
                f"\n   {status}"
            )
            # НЕ зберігаємо target алерти - відправляємо завжди
    
    # Перевіряємо чи треба відправляти "All OK" повідомлення
    should_send = True
    if not downtime_alerts and not target_alerts:
        # Все ОК - перевіряємо коли востаннє відправляли таке повідомлення
        last_ok_alert = sent_alerts.get("_last_all_ok_alert")
        
        if last_ok_alert:
            last_ok_time = datetime.fromisoformat(last_ok_alert)
            hours_since_last = (current_time - last_ok_time).total_seconds() / 3600
            
            if hours_since_last < 4:
                # Менше 4 годин - не відправляємо
                log(f"All OK, but last message was {hours_since_last:.1f}h ago (< 4h) - skipping")
                should_send = False
            else:
                log(f"All OK, last message was {hours_since_last:.1f}h ago - sending")
        else:
            log("All OK - sending first time")
    
    # Відправляємо
    if should_send:
        send_telegram("\n".join(lines))
        
        # Якщо це було "All OK" повідомлення - запам'ятовуємо час
        if not downtime_alerts and not target_alerts:
            sent_alerts["_last_all_ok_alert"] = current_time.isoformat()
    
    # Зберігаємо оновлений список
    with open(sent_alerts_file, "w", encoding="utf-8") as f:
        json.dump({
            "last_reset": current_time.isoformat(),
            "alerts": sent_alerts
        }, f, indent=2)

# ── HTML generation ───────────────────────────────────────────────────────────
def fmt_time(dt):   return dt.strftime("%H:%M") if dt else "—"
def eff_color(pct): return "#22c55e" if pct >= 75 else ("#f59e0b" if pct >= 50 else "#ef4444")

def generate_html(cycles, downtimes, period_from, period_to, timeline_data, conn, excel_targets, mr_data):
    generated  = datetime.now().strftime("%d.%m.%Y %H:%M")
    period_str = f"{fmt_time(period_from)} – {fmt_time(period_to)}"

    def timeline_bar(mname):
        segs      = timeline_data.get(mname, [])
        bars      = ""
        for s in segs:
            if s["w"] <= 0.1:
                continue
            color   = "#22c55e" if s["state"] == "1" else "#ef4444"
            seg_id  = s["id"]
            label   = s["label"].replace("'", "&#39;").replace('"', "&quot;")
            tip     = f'{s["start"]}–{s["end"]} | {label}'
            
            # Додаємо текст всередині сегмента якщо він достатньо широкий
            seg_text = ""
            if s["w"] > 3:  # якщо ширина > 3% показуємо час
                seg_text = f'<span style="position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);font-size:11px;font-weight:600;color:white;white-space:nowrap;pointer-events:none">{s["start"]}</span>'
            
            bars += (
                f'<div class="tl-seg" '
                f'data-id="{seg_id}" '
                f'data-tip="{tip}" '
                f'style="position:absolute;left:{s["x"]:.2f}%;width:{s["w"]:.2f}%;'
                f'height:100%;background:{color};cursor:pointer;'
                f'transition:filter .15s,opacity .15s">{seg_text}</div>'
            )
        total_min = max(round((period_to - period_from).total_seconds() / 60, 2), 0.01)
        ticks     = ""
        for i in range(0, int(total_min) + 1, 15):
            pct       = i / total_min * 100
            tick_time = (period_from + timedelta(minutes=i)).strftime("%H:%M")
            # кожні 30 хв — завжди видно; кожні 15 хв — ховати на mobile
            extra_cls = "" if i % 30 == 0 else " tl-tick-15"
            ticks += (f'<div class="tl-tick{extra_cls}" '
                      f'style="left:{pct:.1f}%">{tick_time}</div>')
        short = mname.split("_")[0] if "_" in mname else mname
        # Таймлайн розтягнутий для 24 годин (мінімум 100px на годину)
        timeline_width = max(2400, total_min * 1.67)  # 1.67px на хвилину
        
        return (
            f'<div class="tl-scroll-wrapper" data-machine="{short}">'
            f'<div class="tl-inner" style="width:{timeline_width}px">'
            f'<div class="tl-wrap" data-machine="{short}" '
            f'style="position:relative;height:44px;background:#f1f5f9">'
            f'{bars}</div>'
            f'<div class="tl-ticks">{ticks}</div>'
            f'</div></div>'
        )

    def history_chart(mname):
        rows_h = load_history(conn, mname)
        if not rows_h:
            return ""
        cid    = mname.replace(" ", "_").replace("-", "_")
        labels = json.dumps([r[0][5:] for r in rows_h])
        eff    = json.dumps([r[1] for r in rows_h])
        run    = json.dumps([r[2] for r in rows_h])
        down   = json.dumps([r[3] for r in rows_h])
        return f'''
        <div class="section-title">📈 7-Day History</div>
        <div style="padding:12px 20px 16px">
          <canvas id="chart_{cid}" height="80"></canvas>
        </div>
        <script>
        (function(){{
          new Chart(document.getElementById("chart_{cid}").getContext("2d"),{{
            type:"bar",
            data:{{labels:{labels},datasets:[
              {{label:"Run (min)",data:{run},backgroundColor:"#22c55e88"}},
              {{label:"Down (min)",data:{down},backgroundColor:"#ef444488"}},
              {{label:"Efficiency %",data:{eff},type:"line",borderColor:"#3b82f6",
               backgroundColor:"transparent",yAxisID:"y2",pointRadius:4,borderWidth:2}}
            ]}},
            options:{{responsive:true,interaction:{{mode:"index"}},
              plugins:{{legend:{{position:"top"}}}},
              scales:{{
                y:{{title:{{display:true,text:"minutes"}}}},
                y2:{{position:"right",min:0,max:100,
                     title:{{display:true,text:"efficiency %"}},
                     grid:{{drawOnChartArea:false}}}}
              }}
            }}
          }});
        }})();
        </script>'''

    def activity_section(c_list, d_list, mname):
        """Об'єднана таблиця циклів та простоїв, відсортована за часом"""
        segs = timeline_data.get(mname, [])
        
        # Функція пошуку ID для циклів
        def find_cycle_ids(cycle):
            from datetime import datetime
            c_start = cycle.get("start")
            if not c_start:  # Якщо start=None (розділений цикл)
                return ""
            c_end   = cycle.get("end") or datetime.now()
            ids = []
            for s in segs:
                if s["state"] != "1":
                    continue
                s_start = datetime.strptime(f"{c_start.strftime('%Y.%m.%d')} {s['start']}", "%Y.%m.%d %H:%M")
                s_end   = datetime.strptime(f"{c_start.strftime('%Y.%m.%d')} {s['end']}", "%Y.%m.%d %H:%M")
                if s_start <= c_end and s_end >= c_start:
                    ids.append(s["id"])
            return " ".join(ids) if ids else ""
        
        # Функція пошуку ID для простоїв
        def find_down_id(down):
            for s in segs:
                if s["state"] == "0" and s["start"] == down["start"].strftime("%H:%M"):
                    return s["id"]
                if s["state"] == "0" and s["start"] <= down["start"].strftime("%H:%M") <= s["end"]:
                    return s["id"]
            return ""
        
        # Об'єднуємо всі події
        events = []
        
        # Додаємо цикли
        for c in c_list:
            # Для розділених циклів (start=None) використовуємо фіктивний час для сортування
            # але не показуємо його користувачу
            cycle_start = c["start"] if c.get("start") else datetime.now()
            
            events.append({
                "type": "cycle",
                "start": cycle_start,  # Для сортування
                "end": c.get("end"),
                "duration": c["duration"],
                "program": c.get("program", "—"),
                "ongoing": c.get("ongoing", False),
                "split_from_schedule": c.get("split_from_schedule", False),  # Позначка
                "ids": find_cycle_ids(c)  # Функція сама обробляє None
            })
        
        # Додаємо простої
        for d in d_list:
            events.append({
                "type": "downtime",
                "start": d["start"],
                "end": d.get("end"),
                "duration": d["duration"],
                "reason": d["reason"],
                "ongoing": d.get("ongoing", False),
                "ids": find_down_id(d)
            })
        
        # Сортуємо за часом
        events.sort(key=lambda e: e["start"])
        
        if not events:
            return '<p class="empty">No activity detected</p>'
        
        # Генеруємо рядки таблиці
        rows_html = ""
        for e in events:
            badge = ' <span class="badge ongoing">ongoing</span>' if e["ongoing"] else ""
            end_s = "…" if not e.get("end") else fmt_time(e["end"])
            
            if e["type"] == "cycle":
                icon = "🟢"
                detail = e["program"]
                row_class = "activity-run"
                # Для розділених циклів показуємо "—"
                if e.get("split_from_schedule"):
                    start_s = "—"
                    end_s = "—"
                else:
                    start_s = fmt_time(e["start"])
                    end_s = "…" if not e.get("end") else fmt_time(e["end"])
            else:
                icon = "🔴"
                detail = e["reason"]
                row_class = "activity-down"
                start_s = fmt_time(e["start"])
                end_s = "…" if not e.get("end") else fmt_time(e["end"])
            
            rows_html += (
                f'<tr class="tl-row {row_class}" data-id="{e["ids"]}">'
                f'<td>{detail}</td>'
                f'<td>{icon}</td>'
                f'<td>{start_s}</td>'
                f'<td>{end_s}{badge}</td>'
                f'<td><strong>{e["duration"]} min</strong></td></tr>'
            )
        
        max_h = "600px" if len(events) > 8 else "400px"
        # Унікальний ID для цього блоку
        scroll_id = f"activity-scroll-{mname.replace('_', '-')}"
        return (
            f'<div class="resizable-section" style="height:{max_h}">'
            f'<div class="table-scroll-x" style="height:100%;overflow:hidden"><div class="scroll-table-wrap" style="height:100%">'
            f'<table class="scroll-table"><thead><tr><th>Details</th><th></th><th>Start</th><th>End</th><th>Duration</th></tr></thead></table>'
            f'<div id="{scroll_id}" class="scroll-tbody-wrap" style="height:calc(100% - 40px);overflow-y:auto">'
            f'<table class="scroll-table"><tbody>{rows_html}</tbody></table>'
            f'</div></div></div></div>'
        )

    def cycles_section(c_list, mname, excel_targets, conn, mr_data):
        """Генерує Target Cycle Time з порівнянням з Excel"""
        if not c_list:
            return ""
        
        # DEBUG: Логуємо скільки targets завантажено
        log(f"cycles_section: machine={mname}, excel_targets count={len(excel_targets)}")
        
        # Витягуємо коротку назву станку (M1, M2 тощо)
        machine_short = mname.split("_")[0] if "_" in mname else mname
        
        # Групуємо цикли по програмах
        by_prog = defaultdict(list)
        for c in c_list:
            prog_name = c["program"] or "—"
            if prog_name not in by_prog:
                by_prog[prog_name] = []
            by_prog[prog_name].append(c)
        
        # Для кожної програми використовуємо нову логіку вибірки
        target_rows = []
        for prog, current_cycles in by_prog.items():
            # Визначаємо операцію
            op_num = get_operation_number(prog)
            
            # Розраховуємо реальну довжину циклу РОЗУМНО
            # (пріоритет: MachiningResult → фільтровані → всі)
            calc_target = calculate_cycle_time_smart(mname, prog, current_cycles, mr_data)
            
            if calc_target is None:
                # Недостатньо даних
                continue
            
            info_text = f"{len(current_cycles)} cycles today"
            
            # Шукаємо Excel Target з урахуванням станку та операції
            excel_target = None
            prog_normalized = normalize_program_name(prog)
            found_for_other_machine = None
            
            # Перебираємо всі ключі в excel_targets
            for (excel_prog, excel_op, excel_machine), time_val in excel_targets.items():
                excel_prog_normalized = normalize_program_name(excel_prog)
                excel_machine_normalized = normalize_program_name(excel_machine)
                machine_normalized = normalize_program_name(machine_short)
                
                # Порівнюємо: програма + операція + станок
                if (prog_normalized == excel_prog_normalized and 
                    op_num == excel_op and 
                    machine_normalized == excel_machine_normalized):
                    excel_target = time_val
                    break
                
                # Зберігаємо якщо знайшли для іншого станку
                if (prog_normalized == excel_prog_normalized and 
                    op_num == excel_op and 
                    machine_normalized != excel_machine_normalized):
                    found_for_other_machine = excel_machine
            
            # Порівняння
            if excel_target:
                diff = round(calc_target - excel_target, 2)  # Округлення до сотих
                diff_pct = round((diff / excel_target) * 100, 2) if excel_target else 0  # Також до сотих
                if diff > 0:
                    comparison = f'<span style="color:#ef4444">+{diff} min (+{diff_pct}%)</span>'
                elif diff < 0:
                    comparison = f'<span style="color:#22c55e">{diff} min ({diff_pct}%)</span>'
                else:
                    comparison = '<span style="color:#64748b">Match</span>'
                excel_col = f'{excel_target} min'
            else:
                # Якщо не знайшли для цього станку, але є для іншого
                if found_for_other_machine:
                    comparison = '<span style="color:#f59e0b">Wrong machine</span>'
                    excel_col = f'No norm for {machine_short}'
                else:
                    comparison = '<span style="color:#94a3b8">No data</span>'
                    excel_col = '—'
            
            target_rows.append(
                f'<tr><td><em>{prog}</em></td>'
                f'<td><span class="badge" style="background:#3b82f6;color:white;padding:2px 8px;border-radius:10px">OP{op_num}</span></td>'
                f'<td>{len(current_cycles)}</td>'
                f'<td><strong>{calc_target} min</strong></td>'
                f'<td>{excel_col}</td>'
                f'<td>{comparison}</td>'
                f'<td style="font-size:0.75rem;color:#64748b">{info_text}</td></tr>'
            )

        return (
            f'<div class="section-title" style="border-top:2px dashed #e2e8f0">🎯 Target Cycle Time</div>'
            f'<div class="table-scroll-x"><table><thead><tr><th>Program</th><th>OP</th><th>Total</th><th>Calculated</th><th>Target</th><th>Difference</th><th>Info</th></tr></thead>'
            f'<tbody>{"".join(target_rows)}</tbody></table></div>'
        )

    machines_html = ""
    for mname in sorted(cycles.keys()):
        c_list     = cycles.get(mname, [])
        d_data     = downtimes.get(mname, {})
        d_list     = d_data.get("downtimes", [])
        total_min  = d_data.get("total_min", 1)
        total_run  = d_data.get("total_run", 0)
        total_down = d_data.get("total_down", 0)
        eff        = round(total_run / total_min * 100) if total_min else 0
        short_name = mname.split("_")[0] if "_" in mname else mname

        segs_for_down = timeline_data.get(mname, [])

        def find_down_seg_id(d):
            for s in segs_for_down:
                if s["state"] == "0" and s["start"] == d["start"].strftime("%H:%M"):
                    return s["id"]
                if s["state"] == "0" and s["start"] <= d["start"].strftime("%H:%M") <= s["end"]:
                    return s["id"]
            return ""

        down_rows = "".join(
            f'<tr class="tl-row" data-id="{find_down_seg_id(d)}">'
            f'<td>{fmt_time(d["start"])}</td>'
            f'<td>{"…" if not d.get("end") else fmt_time(d["end"])}'
            f'{"<span class=\"badge ongoing\">ongoing</span>" if d.get("ongoing") else ""}</td>'
            f'<td><strong>{d["duration"]} min</strong></td><td>{d["reason"]}</td></tr>'
            for d in d_list)
        down_max_h = "200px" if len(d_list) > 5 else "auto"
        down_table = (
            f'<div class="table-scroll-x"><div class="scroll-table-wrap">'
            f'<table class="scroll-table"><thead><tr><th>Start</th><th>End</th><th>Duration</th><th>Reason</th></tr></thead></table>'
            f'<div class="scroll-tbody-wrap" style="max-height:{down_max_h};overflow-y:auto">'
            f'<table class="scroll-table"><tbody>{down_rows}</tbody></table>'
            f'</div></div></div>'
            if d_list else '<p class="empty">No downtime detected</p>')

        machines_html += f"""
        <div class="machine-card">
          <div class="machine-header">
            <div class="machine-title">
              <span class="machine-id">{short_name}</span>
              <span class="machine-full">{mname}</span>
            </div>
            <div class="eff-badge" style="background:{eff_color(eff)}">
              Efficiency: {eff}%
              <span class="eff-detail">({total_run} / {total_min} min)</span>
            </div>
          </div>
          <div class="section-title">⏱ Timeline</div>
          <div style="padding:10px 20px 4px">{timeline_bar(mname)}</div>
          <div class="section-title">📋 Activity Log — {len(c_list)} cycles, {len(d_list)} downtimes ({total_down} min)</div>
          <div style="padding:0 0 4px">{activity_section(c_list, d_list, mname)}</div>
          {cycles_section(c_list, mname, excel_targets, conn, mr_data)}
          {history_chart(mname)}
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Machine Report — {generated}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#1a1a2e;font-size:15px;margin:0}}

  /* ── Header ── */
  .header{{background:#1a1a2e;color:white;padding:16px 20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}}
  .header h1{{font-size:1.2rem;font-weight:600}}
  .header .meta{{font-size:.8rem;opacity:.7;text-align:right}}

  /* ── Layout ── */
  .container{{max-width:1100px;margin:16px auto;padding:0 12px}}

  /* ── Machine card ── */
  .machine-card{{background:white;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:20px;overflow:hidden}}
  .machine-header{{background:#1e293b;color:white;padding:14px 16px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}}
  .machine-title{{display:flex;flex-direction:column;gap:2px}}
  .machine-id{{font-size:1.1rem;font-weight:700}}
  .machine-full{{font-size:.7rem;opacity:.6;word-break:break-all}}
  .eff-badge{{padding:5px 12px;border-radius:20px;font-weight:600;font-size:.85rem;color:white;white-space:nowrap}}
  .eff-detail{{font-size:.72rem;font-weight:400;opacity:.85;margin-left:4px}}

  /* ── Section title ── */
  .section-title{{padding:10px 16px 5px;font-weight:600;font-size:.8rem;color:#475569;border-top:1px solid #f1f5f9;text-transform:uppercase;letter-spacing:.03em}}

  /* ── Tables (desktop) ── */
  table{{width:100%;border-collapse:collapse;font-size:.85rem}}
  th{{background:#f8fafc;padding:8px 12px;text-align:left;font-weight:600;color:#64748b;border-bottom:1px solid #e2e8f0;white-space:nowrap}}
  td{{padding:8px 12px;border-bottom:1px solid #f1f5f9;word-break:break-word}}
  tr:last-child td{{border-bottom:none}}
  tr:hover td{{background:#f8fafc}}
  .tl-row{{cursor:pointer;transition:background-color 0.2s ease}}
  .tl-row:hover td{{background:#e0f2fe!important}}
  .activity-run td{{background:#f0fdf4}}
  .activity-down td{{background:#fef2f2}}
  .activity-run:hover td{{background:#dcfce7!important}}
  .activity-down:hover td{{background:#fee2e2!important}}

  /* ── Scrollable tables ── */
  .scroll-table-wrap{{width:100%;border-bottom:1px solid #e2e8f0}}
  .scroll-table{{width:100%;border-collapse:collapse;table-layout:fixed;font-size:.85rem}}
  .scroll-table th,.scroll-table td{{padding:8px 12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
  .scroll-table thead th{{background:#f8fafc;font-weight:600;color:#64748b;border-bottom:1px solid #e2e8f0;position:sticky;top:0;z-index:1}}
  .scroll-table tbody tr:last-child td{{border-bottom:none}}
  .scroll-table tbody tr:hover td{{background:#f8fafc}}
  .scroll-tbody-wrap{{display:block}}
  .scroll-tbody-wrap::-webkit-scrollbar{{width:5px}}
  .scroll-tbody-wrap::-webkit-scrollbar-track{{background:#f1f5f9}}
  .scroll-tbody-wrap::-webkit-scrollbar-thumb{{background:#cbd5e1;border-radius:3px}}
  
  /* ── Resizable Activity Log ── */
  .resizable-section{{
    position:relative;
    overflow:hidden;
    border:1px solid #e2e8f0;
    border-radius:6px;
    background:#fff;
  }}
  
  /* ── Resize Handle (як у Windows) ── */
  .resize-handle{{
    position:absolute;
    bottom:0;
    right:0;
    width:16px;
    height:16px;
    cursor:nwse-resize;
    background:linear-gradient(135deg, transparent 0%, transparent 40%, #cbd5e1 40%, #cbd5e1 45%, transparent 45%, transparent 55%, #cbd5e1 55%, #cbd5e1 60%, transparent 60%, transparent 70%, #cbd5e1 70%, #cbd5e1 75%, transparent 75%);
    z-index:10;
  }}
  .resize-handle:hover{{
    background:linear-gradient(135deg, transparent 0%, transparent 40%, #94a3b8 40%, #94a3b8 45%, transparent 45%, transparent 55%, #94a3b8 55%, #94a3b8 60%, transparent 60%, transparent 70%, #94a3b8 70%, #94a3b8 75%, transparent 75%);
  }}

  /* ── Timeline ── */
  .tl-scroll-wrapper{{overflow-x:auto;overflow-y:hidden;position:relative;margin:0 -12px;padding:0 12px}}
  .tl-scroll-wrapper::-webkit-scrollbar{{height:8px}}
  .tl-scroll-wrapper::-webkit-scrollbar-track{{background:#f1f5f9}}
  .tl-scroll-wrapper::-webkit-scrollbar-thumb{{background:#cbd5e1;border-radius:4px}}
  .tl-inner{{position:relative}}
  .tl-wrap{{overflow:hidden!important;border-radius:4px;position:relative;user-select:none}}
  .tl-ticks{{position:relative;height:20px;user-select:none}}
  .tl-tick{{position:absolute;top:24px;font-size:10px;color:#94a3b8;transform:translateX(-50%)}}

  /* ── Legend ── */
  .legend{{display:flex;gap:14px;padding:4px 16px 10px;font-size:.78rem;flex-wrap:wrap}}
  .legend span{{display:flex;align-items:center;gap:4px}}
  .dot{{width:11px;height:11px;border-radius:2px;display:inline-block;flex-shrink:0}}

  /* ── Misc ── */
  .badge{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:.72rem;font-weight:600;margin-left:3px}}
  .badge.ongoing{{background:#dbeafe;color:#1d4ed8}}
  .empty{{padding:12px 16px;color:#94a3b8;font-style:italic;font-size:.85rem}}
  .footer{{text-align:center;color:#94a3b8;font-size:.75rem;padding:16px;word-break:break-all}}

  /* ── Tooltip ── */
  #tl-tooltip{{
    position:fixed;pointer-events:none;z-index:9999;
    background:#1e293b;color:white;
    padding:6px 10px;border-radius:6px;font-size:.78rem;
    box-shadow:0 4px 12px rgba(0,0,0,.3);
    display:none;max-width:240px;white-space:normal;line-height:1.4
  }}

  /* ── Highlight ── */
  .tl-seg.dim{{opacity:.25}}
  .tl-seg.highlight{{filter:brightness(1.25);outline:2px solid white;z-index:2;position:relative}}
  .tl-row{{transition:background-color 0.2s ease}}
  .tl-row.highlight td{{background:#fde047!important;font-weight:600}}
  .tl-row:hover td{{background:#f0f9ff!important}}

  /* ══════════════════════════════════════════
     MOBILE  ≤ 600px
  ══════════════════════════════════════════ */
  @media(max-width:600px){{
    body{{font-size:14px}}
    .header{{padding:12px 14px}}
    .header h1{{font-size:1rem}}
    .container{{padding:0 8px;margin:10px auto}}
    .machine-card{{border-radius:8px;margin-bottom:14px}}
    .machine-header{{padding:12px 12px}}
    .machine-id{{font-size:1rem}}
    .eff-badge{{font-size:.78rem;padding:4px 10px}}
    .eff-detail{{display:none}}          /* прибираємо (x/y min) на малих екранах */
    .section-title{{padding:8px 12px 4px;font-size:.75rem}}

    /* Таблиці — горизонтальний скрол замість truncate */
    .table-scroll-x{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
    table,.scroll-table{{font-size:.78rem;min-width:360px}}
    th,td,.scroll-table th,.scroll-table td{{padding:7px 9px}}

    /* Timeline — тікі тільки кожні 30 хв */
    .tl-tick-15{{display:none}}
    .tl-ticks{{height:16px}}
    .tl-tick{{font-size:8px}}

    /* Tooltip — завжди знизу екрану на mobile */
    #tl-tooltip{{
      position:fixed;bottom:12px;left:50%;transform:translateX(-50%);
      top:auto!important;max-width:92vw;text-align:center
    }}

    .legend{{padding:4px 12px 8px;font-size:.74rem;gap:10px}}
    .footer{{font-size:.7rem;padding:12px}}
  }}
</style>
</head>
<body>
<div class="header">
  <h1>📊 Machine Report</h1>
  <div class="meta">
    Period: {period_str}<br>
    Generated: {generated}<br>
    <span style="font-size: 12px; color: #64748b;">⚡ Auto-updates hourly (8-20, Mon-Fri)</span>
  </div>
</div>
<div class="container">
  <div class="legend">
    <span><span class="dot" style="background:#22c55e"></span>Running</span>
    <span><span class="dot" style="background:#ef4444"></span>Downtime</span>
  </div>
  {machines_html}
</div>
<div class="footer">Source: {CSV_FILE} &nbsp;|&nbsp; DB: {DB_FILE}</div>
<div id="tl-tooltip"></div>
<script>
(function(){{
  var tip = document.getElementById("tl-tooltip");

  // ── Tooltip + highlight при наведенні на сегмент таймлайну ────────────────
  document.querySelectorAll(".tl-seg").forEach(function(seg){{
    seg.addEventListener("mouseenter", function(e){{
      var id  = seg.dataset.id;
      var txt = seg.dataset.tip;
      tip.textContent = txt;
      tip.style.display = "block";
      var wrap = seg.closest(".tl-wrap");
      wrap.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.add("dim"); }});
      seg.classList.remove("dim");
      seg.classList.add("highlight");
      // Шукаємо рядки що містять цей ID (може бути кілька ID через пробіл)
      if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids = r.dataset.id ? r.dataset.id.split(' ') : [];
        if(ids.indexOf(id) !== -1) r.classList.add("highlight");
      }});
    }});
    seg.addEventListener("mousemove", function(e){{
      tip.style.left = (e.clientX + 14) + "px";
      tip.style.top  = (e.clientY - 32) + "px";
    }});
    seg.addEventListener("mouseleave", function(){{
      tip.style.display = "none";
      var wrap = seg.closest(".tl-wrap");
      wrap.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.remove("dim","highlight"); }});
      var id = seg.dataset.id;
      if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids = r.dataset.id ? r.dataset.id.split(' ') : [];
        if(ids.indexOf(id) !== -1) r.classList.remove("highlight");
      }});
    }});
    
    // ── Клік на сегмент — скрол до рядка в центр ──
    seg.addEventListener("click", function(e){{
      var id = seg.dataset.id;
      if(!id) return;
      
      // Шукаємо рядок що містить цей ID
      var targetRow = null;
      document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids = r.dataset.id ? r.dataset.id.split(' ') : [];
        if(ids.indexOf(id) !== -1) targetRow = r;
      }});
      
      if(!targetRow) return;
      
      // Знаходимо scroll container (може бути .scroll-tbody-wrap або весь документ)
      var scrollContainer = targetRow.closest(".scroll-tbody-wrap");
      
      if(scrollContainer){{
        // Скрол всередині таблиці
        var rowRect = targetRow.getBoundingClientRect();
        var containerRect = scrollContainer.getBoundingClientRect();
        var targetScroll = scrollContainer.scrollTop + (rowRect.top - containerRect.top) - (containerRect.height / 2) + (rowRect.height / 2);
        
        scrollContainer.scrollTo({{
          top: targetScroll,
          behavior: 'smooth'
        }});
      }} else {{
        // Скрол всієї сторінки
        var rowTop = targetRow.getBoundingClientRect().top + window.scrollY;
        var viewportCenter = window.innerHeight / 2;
        
        window.scrollTo({{
          top: rowTop - viewportCenter + (targetRow.offsetHeight / 2),
          behavior: 'smooth'
        }});
      }}
      
      // Короткий flash ефект для підсвітки
      targetRow.classList.add("highlight");
      setTimeout(function(){{
        targetRow.classList.remove("highlight");
      }}, 1500);
    }});
  }});

  // ── Highlight сегменту при наведенні на рядок таблиці ─────────────────────
  document.querySelectorAll(".tl-row").forEach(function(row){{
    row.addEventListener("mouseenter", function(){{
      var ids = row.dataset.id ? row.dataset.id.split(' ') : [];
      if(ids.length === 0) return;
      
      var wrap = null;
      var segs = [];
      ids.forEach(function(id){{
        var seg = document.querySelector('.tl-seg[data-id="'+id+'"]');
        if(seg){{
          if(!wrap) wrap = seg.closest(".tl-wrap");
          segs.push(seg);
        }}
      }});
      
      if(segs.length === 0) return;
      
      wrap.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.add("dim"); }});
      segs.forEach(function(seg){{
        seg.classList.remove("dim");
        seg.classList.add("highlight");
      }});
      
      // Показуємо tooltip для першого сегмента
      tip.textContent = segs[0].dataset.tip;
      tip.style.display = "block";
      tip.style.left = (row.getBoundingClientRect().right + 10) + "px";
      tip.style.top  = (row.getBoundingClientRect().top + window.scrollY) + "px";
    }});
    row.addEventListener("mouseleave", function(){{
      var ids = row.dataset.id ? row.dataset.id.split(' ') : [];
      if(ids.length === 0) return;
      
      var wrap = null;
      ids.forEach(function(id){{
        var seg = document.querySelector('.tl-seg[data-id="'+id+'"]');
        if(seg){{
          if(!wrap) wrap = seg.closest(".tl-wrap");
          wrap.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.remove("dim","highlight"); }});
        }}
      }});
      tip.style.display = "none";
    }});
    
    // ── Клік на рядок — скрол до сегмента в центр ──
    row.addEventListener("click", function(){{
      var ids = row.dataset.id ? row.dataset.id.split(' ') : [];
      if(ids.length === 0) return;
      
      var firstSeg = document.querySelector('.tl-seg[data-id="'+ids[0]+'"]');
      if(!firstSeg) return;
      
      var wrapper = firstSeg.closest(".tl-scroll-wrapper");
      if(!wrapper) return;
      
      // Позиція сегмента відносно wrapper
      var segRect = firstSeg.getBoundingClientRect();
      var wrapRect = wrapper.getBoundingClientRect();
      
      // Поточна позиція скролу + відстань до сегмента - половина ширини wrapper
      var targetScroll = wrapper.scrollLeft + (segRect.left - wrapRect.left) - (wrapRect.width / 2) + (segRect.width / 2);
      
      // Плавний скрол
      wrapper.scrollTo({{
        left: targetScroll,
        behavior: 'smooth'
      }});
    }});
  }});


  // ── Touch support for mobile ──────────────────────────────────────────
  document.querySelectorAll(".tl-seg").forEach(function(seg){{
    var touchStartTime;
    var touchMoved = false;
    
    seg.addEventListener("touchstart", function(e){{
      touchStartTime = Date.now();
      touchMoved = false;
      e.preventDefault();
      tip.textContent = seg.dataset.tip;
      tip.style.display = "block";
      var wrap = seg.closest(".tl-wrap");
      wrap.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.add("dim"); }});
      seg.classList.remove("dim"); seg.classList.add("highlight");
      var id = seg.dataset.id;
      if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids = r.dataset.id ? r.dataset.id.split(' ') : [];
        if(ids.indexOf(id) !== -1) r.classList.add("highlight");
      }});
    }}, {{passive:false}});
    
    seg.addEventListener("touchmove", function(){{
      touchMoved = true;
    }});
    
    seg.addEventListener("touchend", function(){{
      var touchDuration = Date.now() - touchStartTime;
      var id = seg.dataset.id;
      
      // Якщо це був швидкий тап (не drag), робимо скрол до рядка
      if(!touchMoved && touchDuration < 300 && id){{
        var targetRow = null;
        document.querySelectorAll('.tl-row').forEach(function(r){{
          var ids = r.dataset.id ? r.dataset.id.split(' ') : [];
          if(ids.indexOf(id) !== -1) targetRow = r;
        }});
        
        if(targetRow){{
          var scrollContainer = targetRow.closest(".scroll-tbody-wrap");
          
          if(scrollContainer){{
            var rowRect = targetRow.getBoundingClientRect();
            var containerRect = scrollContainer.getBoundingClientRect();
            var targetScroll = scrollContainer.scrollTop + (rowRect.top - containerRect.top) - (containerRect.height / 2) + (rowRect.height / 2);
            
            scrollContainer.scrollTo({{
              top: targetScroll,
              behavior: 'smooth'
            }});
          }} else {{
            var rowTop = targetRow.getBoundingClientRect().top + window.scrollY;
            var viewportCenter = window.innerHeight / 2;
            
            window.scrollTo({{
              top: rowTop - viewportCenter + (targetRow.offsetHeight / 2),
              behavior: 'smooth'
            }});
          }}
          
          targetRow.classList.add("highlight");
          setTimeout(function(){{
            targetRow.classList.remove("highlight");
          }}, 1500);
        }}
      }}
      
      setTimeout(function(){{
        tip.style.display = "none";
        var wrap = seg.closest(".tl-wrap");
        wrap.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.remove("dim","highlight"); }});
        if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
          var ids = r.dataset.id ? r.dataset.id.split(' ') : [];
          if(ids.indexOf(id) !== -1) r.classList.remove("highlight");
        }});
      }}, 1400);
    }});
  }});


  // Auto-scroll до останньої години при завантаженні
  document.querySelectorAll(".tl-scroll-wrapper").forEach(function(wrapper) {{
    var inner = wrapper.querySelector(".tl-inner");
    if (inner) {{
      // Показуємо останню годину (60 хвилин)
      // При 1.67px на хвилину: 60 хв = 100px
      var lastHourWidth = 100;  // приблизно 60 хвилин
      var scrollLeft = inner.offsetWidth - wrapper.offsetWidth;
      // Якщо таймлайн довший ніж 1 година, скролимо до останньої години
      if (scrollLeft > lastHourWidth) {{
        scrollLeft = Math.max(0, scrollLeft);
      }}
      wrapper.scrollLeft = scrollLeft;
    }}
  }});

  // ── Drag to scroll для таймлайну ──────────────────────────────────────
  document.querySelectorAll(".tl-scroll-wrapper").forEach(function(wrapper) {{
    var isDown = false;
    var startX;
    var scrollLeftStart;
    var isDragging = false;

    var tlWrap = wrapper.querySelector(".tl-wrap");
    if (!tlWrap) return;

    tlWrap.addEventListener("mousedown", function(e) {{
      isDown = true;
      isDragging = false;
      startX = e.pageX;
      scrollLeftStart = wrapper.scrollLeft;
      tlWrap.style.cursor = "grabbing";
      wrapper.style.cursor = "grabbing";
    }});

    document.addEventListener("mousemove", function(e) {{
      if (!isDown) return;
      e.preventDefault();
      isDragging = true;
      var x = e.pageX;
      var walk = (startX - x) * 2; // швидкість скролу
      wrapper.scrollLeft = scrollLeftStart + walk;
    }});

    document.addEventListener("mouseup", function() {{
      if (isDown) {{
        isDown = false;
        tlWrap.style.cursor = "grab";
        wrapper.style.cursor = "";
        // Даємо час для спрацювання click після невеликого руху
        setTimeout(function() {{
          isDragging = false;
        }}, 50);
      }}
    }});

    // Блокуємо click events якщо був drag
    tlWrap.addEventListener("click", function(e) {{
      if (isDragging) {{
        e.stopPropagation();
        e.preventDefault();
      }}
    }}, true);

    // Touch events для mobile
    var touchStartX;
    var touchScrollStart;
    var isTouchDragging = false;

    tlWrap.addEventListener("touchstart", function(e) {{
      touchStartX = e.touches[0].pageX;
      touchScrollStart = wrapper.scrollLeft;
      isTouchDragging = false;
    }}, {{passive: true}});

    tlWrap.addEventListener("touchmove", function(e) {{
      if (!touchStartX) return;
      isTouchDragging = true;
      var touchX = e.touches[0].pageX;
      var walk = (touchStartX - touchX) * 2;
      wrapper.scrollLeft = touchScrollStart + walk;
    }}, {{passive: true}});

    tlWrap.addEventListener("touchend", function() {{
      touchStartX = null;
      setTimeout(function() {{
        isTouchDragging = false;
      }}, 50);
    }});

    // Встановлюємо cursor
    tlWrap.style.cursor = "grab";
  }});
}})();

// ── Activity Log Resize Handle ──
(function() {{
  var resizables = document.querySelectorAll('.resizable-section');
  
  resizables.forEach(function(container) {{
    var isResizing = false;
    var startY = 0;
    var startHeight = 0;
    
    // Створюємо ручку resize
    var handle = document.createElement('div');
    handle.className = 'resize-handle';
    handle.innerHTML = '⋮⋮⋮';
    container.appendChild(handle);
    
    handle.addEventListener('mousedown', function(e) {{
      isResizing = true;
      startY = e.clientY;
      startHeight = container.offsetHeight;
      e.preventDefault();
      document.body.style.userSelect = 'none';
      document.body.style.cursor = 'ns-resize';
    }});
    
    document.addEventListener('mousemove', function(e) {{
      if (!isResizing) return;
      
      var delta = e.clientY - startY;
      var newHeight = startHeight + delta;
      
      // Обмеження висоти
      if (newHeight < 200) newHeight = 200;
      if (newHeight > 1200) newHeight = 1200;
      
      container.style.height = newHeight + 'px';
    }});
    
    document.addEventListener('mouseup', function() {{
      if (isResizing) {{
        isResizing = false;
        document.body.style.userSelect = '';
        document.body.style.cursor = '';
      }}
    }});
  }});

}})();
</script>
</body>
</html>"""
# =============================================================================
def main():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    # Step 1+2 — download both files
    log("=" * 60)
    log("FACTORY MONITOR START")
    log("=" * 60)
    if not download_both_files():
        log("Download failed — aborting.")
        sys.exit(1)

    # Step 3 — analyze
    log("── Step 3: Analyzing data ──")
    rows = load_csv()
    log(f"Rows loaded: {len(rows)}")

    filtered, period_from, period_to = filter_last_hours(rows, HOURS_BACK)
    date_str = period_to.strftime("%Y-%m-%d")
    log(f"Period: {period_from.strftime('%H:%M')} – {period_to.strftime('%H:%M')} ({len(filtered)} rows)")

    cycles        = analyze_cycles(filtered)
    
    # Split schedule program cycles if machining_results.csv is available
    cycles, mr_data = split_schedule_cycles(cycles, MACHINING_RESULT)
    
    downtimes     = analyze_downtime(filtered)
    timeline_data = build_timeline_data(filtered, period_from, period_to)

    conn = init_db()
    save_to_db(conn, date_str, cycles, downtimes)
    log("History saved to DB")

    # Step 3.5 — load Excel target times (перед алертами!)
    log("── Step 3.5: Loading Excel target times ──")
    excel_targets = load_target_times()

    # Step 3.6 — check and send alerts
    check_and_alert(downtimes, period_to, cycles, excel_targets, mr_data)

    # Step 4 — report
    log("── Step 4: Generating report ──")
    try:
        html = generate_html(cycles, downtimes, period_from, period_to, timeline_data, conn, excel_targets, mr_data)
    except Exception as e:
        log(f"✗ Error generating HTML: {e}")
        import traceback
        log(traceback.format_exc())
        raise
    
    conn.close()

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"Report saved: {OUTPUT_HTML}")

    # Step 5 — publish to GitHub Pages
    log("── Step 5: Publishing to GitHub Pages ──")
    publish_to_github(html)
    
    log("=" * 60)
    log("FACTORY MONITOR COMPLETE")
    log("=" * 60)

if __name__ == "__main__":
    main()
